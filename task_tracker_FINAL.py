#!/usr/bin/env python3
"""
Task Tracker - FINAL VERSION with Notes Support
Reads Claim Sheet + Model Decomp, generates dashboards with tabs
Detects silent people, personalizes messages with first name only
Tracks notes on why people are silent
"""

import json
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from collections import defaultdict
from difflib import SequenceMatcher

# Notes storage file
NOTES_FILE = Path.cwd() / "silent_tasker_notes.json"

def load_notes():
    """Load existing notes about silent taskers"""
    if NOTES_FILE.exists():
        with open(NOTES_FILE) as f:
            return json.load(f)
    return {}

def save_notes(notes):
    """Save notes to file"""
    with open(NOTES_FILE, 'w') as f:
        json.dump(notes, f, indent=2)

def load_slack_users():
    """Load Slack users from cache"""
    cache_file = Path.cwd() / "slack_users_tsip_contributors.json"
    if not cache_file.exists():
        print(f"ERROR: slack_users_tsip_contributors.json not found")
        return []
    with open(cache_file) as f:
        return json.load(f).get("users", [])

def find_latest_excel():
    """Find most recent Excel file"""
    excel_files = list(Path.cwd().glob("*.xlsx"))
    if not excel_files:
        print(f"ERROR: No Excel files found")
        return None
    return max(excel_files, key=lambda p: p.stat().st_mtime)

def extract_first_name(full_name):
    """Extract first name from 'First Last' format"""
    if not full_name:
        return ""
    parts = str(full_name).strip().split()
    return parts[0] if parts else ""

def read_claim_sheet(excel_path):
    """Read Claim Sheet"""
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb["Claim Sheet"]
    
    headers = {}
    for col_num, cell in enumerate(sheet[1], 1):
        if cell.value:
            headers[cell.value] = col_num
    
    tasker_col = headers.get("Claimed By")
    date_col = headers.get("Date fixed")
    completion_col = headers.get("Task Completed")
    
    if not all([tasker_col, date_col, completion_col]):
        return None, []
    
    tasker_by_date = defaultdict(lambda: defaultdict(int))
    all_taskers = set()
    
    for row_num in range(2, sheet.max_row + 1):
        tasker_name = sheet.cell(row_num, tasker_col).value
        task_date = sheet.cell(row_num, date_col).value
        is_completed = sheet.cell(row_num, completion_col).value
        
        if not all([tasker_name, task_date, is_completed]):
            continue
        
        tasker_name = str(tasker_name).strip()
        all_taskers.add(tasker_name)
        
        if isinstance(task_date, datetime):
            task_date = task_date.date()
        elif isinstance(task_date, str):
            try:
                task_date = datetime.strptime(task_date, "%Y-%m-%d").date()
            except:
                continue
        
        tasker_by_date[tasker_name][task_date] += 1
    
    return tasker_by_date, sorted(all_taskers)

def read_decomp_sheet(excel_path):
    """Read Model Decomp Sheet"""
    wb = openpyxl.load_workbook(excel_path)
    if "Model Decomp" not in wb.sheetnames:
        return None, []
    
    sheet = wb["Model Decomp"]
    
    headers = {}
    for col_num, cell in enumerate(sheet[1], 1):
        if cell.value:
            headers[cell.value] = col_num
    
    tasker_col = headers.get("Claimed By")
    prompt_date_col = headers.get("Date Prompt Com")
    completion_date_col = headers.get("Date Completed")
    
    if not tasker_col:
        return None, []
    
    tasker_by_date = defaultdict(lambda: defaultdict(int))
    all_taskers = set()
    
    for row_num in range(2, sheet.max_row + 1):
        tasker_name = sheet.cell(row_num, tasker_col).value
        
        if not tasker_name:
            continue
        
        tasker_name = str(tasker_name).strip()
        all_taskers.add(tasker_name)
        
        latest_date = None
        
        if prompt_date_col:
            prompt_date = sheet.cell(row_num, prompt_date_col).value
            if prompt_date:
                if isinstance(prompt_date, datetime):
                    latest_date = prompt_date.date()
                elif isinstance(prompt_date, str):
                    try:
                        latest_date = datetime.strptime(prompt_date, "%Y-%m-%d").date()
                    except:
                        pass
        
        if completion_date_col:
            completion_date = sheet.cell(row_num, completion_date_col).value
            if completion_date:
                if isinstance(completion_date, datetime):
                    comp_date = completion_date.date()
                elif isinstance(completion_date, str):
                    try:
                        comp_date = datetime.strptime(completion_date, "%Y-%m-%d").date()
                    except:
                        comp_date = None
                else:
                    comp_date = None
                
                if comp_date and (latest_date is None or comp_date > latest_date):
                    latest_date = comp_date
        
        if latest_date:
            tasker_by_date[tasker_name][latest_date] += 1
    
    return tasker_by_date, sorted(all_taskers)

def get_silent_taskers(tasker_by_date, all_taskers):
    """Find silent people (48+ hours)"""
    silent_threshold = timedelta(hours=48)
    check_time = datetime.now()
    silent_list = []
    
    for tasker in all_taskers:
        dates = tasker_by_date.get(tasker, {})
        if not dates:
            continue
        
        last_task_date = max(dates.keys())
        time_since_task = check_time - datetime.combine(last_task_date, datetime.min.time())
        
        if time_since_task > silent_threshold:
            first_name = extract_first_name(tasker)
            silent_list.append({
                "name": tasker,
                "first_name": first_name,
                "last_task_date": last_task_date.isoformat(),
                "hours_silent": int(time_since_task.total_seconds() / 3600),
                "days_silent": round(time_since_task.total_seconds() / 86400, 1)
            })
    
    return sorted(silent_list, key=lambda x: x["hours_silent"], reverse=True)

def fuzzy_match_user(sheet_name, slack_users):
    """Fuzzy match sheet name to Slack user"""
    sheet_name_lower = sheet_name.lower()
    parts = sheet_name_lower.split()
    
    scores = []
    
    for user in slack_users:
        username = user.get("username", "").lower()
        real_name = user.get("real_name", "").lower()
        
        score = 0
        
        for part in parts:
            if len(part) > 1:
                if part in username or part in real_name:
                    score += 30
        
        seq_match = SequenceMatcher(None, sheet_name_lower, real_name).ratio()
        score += seq_match * 40
        
        seq_match_user = SequenceMatcher(None, sheet_name_lower, username).ratio()
        score += seq_match_user * 20
        
        if score > 0:
            scores.append({
                "user": user,
                "score": score,
                "username": username,
                "real_name": real_name
            })
    
    if not scores:
        return "none", None
    
    scores.sort(key=lambda x: x["score"], reverse=True)
    best_score = scores[0]["score"]
    
    if best_score > 90:
        return "clear", scores[0]["user"]
    
    candidates = [s for s in scores if s["score"] > 70]
    if len(candidates) > 1:
        return "ambiguous", candidates
    
    if best_score > 60:
        return "clear", scores[0]["user"]
    
    return "none", None

def generate_dashboards(claim_data, claim_taskers, decomp_data, decomp_taskers, slack_users):
    """Generate dashboard and approval UI with notes"""
    
    notes = load_notes()
    
    # Get all dates from both sheets
    all_dates = set()
    for dates_dict in claim_data.values():
        all_dates.update(dates_dict.keys())
    for dates_dict in decomp_data.values():
        all_dates.update(dates_dict.keys())
    
    if not all_dates:
        print("No task data found")
        return
    
    sorted_dates = sorted(all_dates)
    date_headers = [d.strftime("%a %m/%d") for d in sorted_dates]
    
    # Get silent taskers
    claim_silent = get_silent_taskers(claim_data, claim_taskers)
    decomp_silent = get_silent_taskers(decomp_data, decomp_taskers)
    
    # === DASHBOARD with NOTES ===
    html_dashboard = f"""<!DOCTYPE html>
<html>
<head>
    <title>Task Tracker Dashboard</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 32px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        h1 {{ font-size: 28px; font-weight: 600; color: #1a1a1a; margin-bottom: 8px; }}
        .meta {{ color: #666; font-size: 14px; margin-bottom: 24px; }}
        .tabs {{ display: flex; border-bottom: 2px solid #e0e0e0; margin-bottom: 24px; gap: 0; }}
        .tab-button {{ padding: 12px 20px; background: none; border: none; cursor: pointer; font-size: 15px; font-weight: 500; color: #666; border-bottom: 3px solid transparent; margin-bottom: -2px; }}
        .tab-button.active {{ color: #1a1a1a; border-bottom-color: #3b82f6; }}
        .tab-button:hover {{ color: #1a1a1a; }}
        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; }}
        .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 16px; margin-bottom: 24px; }}
        .stat-card {{ background: #f9f9f9; padding: 16px; border-radius: 6px; border-left: 4px solid #3b82f6; }}
        .stat-label {{ color: #666; font-size: 12px; text-transform: uppercase; margin-bottom: 6px; }}
        .stat-value {{ font-size: 24px; font-weight: 600; color: #1a1a1a; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #e0e0e0; }}
        th {{ background: #f9f9f9; font-weight: 600; color: #1a1a1a; }}
        th.date {{ text-align: center; width: 70px; }}
        td.date {{ text-align: center; font-weight: 500; color: #0066cc; }}
        tr:hover {{ background: #fafafa; }}
        .silent {{ background: #fffbea; }}
        .active {{ background: #f0fdf4; }}
        .badge {{ display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: 500; }}
        .badge.warning {{ background: #fcd34d; color: #7c2d12; }}
        .badge.success {{ background: #bbf7d0; color: #166534; }}
        .notes-column {{ max-width: 200px; font-size: 12px; color: #666; }}
        .note-tag {{ background: #e3f2fd; color: #1565c0; padding: 2px 6px; border-radius: 3px; font-size: 11px; margin-top: 4px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Task Tracker Dashboard</h1>
        <div class="meta">Generated: {datetime.now().strftime('%A, %B %d, %Y at %I:%M %p')} EST</div>
        
        <div class="tabs">
            <button class="tab-button active" onclick="openTab(event, 'claim')">Claim Sheet Activity</button>
            <button class="tab-button" onclick="openTab(event, 'decomp')">Decomp Progress</button>
            <button class="tab-button" onclick="openTab(event, 'silent')">Silent Taskers & Notes</button>
            <button class="tab-button" onclick="openTab(event, 'historic')">Historic Data</button>
        </div>
        
        <!-- CLAIM SHEET TAB -->
        <div id="claim" class="tab-content active">
            <div class="stats">
                <div class="stat-card">
                    <div class="stat-label">Total Taskers</div>
                    <div class="stat-value">{len(claim_taskers)}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Active</div>
                    <div class="stat-value">{len(claim_taskers) - len(claim_silent)}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Silent (48+hrs)</div>
                    <div class="stat-value" style="color: #dc2626;">{len(claim_silent)}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Total Tasks</div>
                    <div class="stat-value">{sum(sum(dates.values()) for dates in claim_data.values())}</div>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Tasker Name</th>
"""
    
    for date_header in date_headers:
        html_dashboard += f"                        <th class='date'>{date_header}</th>\n"
    
    html_dashboard += """                        <th class='date'>Total</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    silent_names = {s["name"] for s in claim_silent}
    
    for tasker in sorted(claim_taskers):
        row_class = "silent" if tasker in silent_names else "active"
        html_dashboard += f"                    <tr class='{row_class}'>\n"
        html_dashboard += f"                        <td>{tasker}</td>\n"
        
        daily_counts = []
        for date in sorted_dates:
            count = claim_data[tasker].get(date, 0)
            daily_counts.append(count)
            html_dashboard += f"                        <td class='date'>{count if count > 0 else '-'}</td>\n"
        
        total = sum(daily_counts)
        html_dashboard += f"                        <td class='date'>{total}</td>\n"
        
        if tasker in silent_names:
            html_dashboard += f"                        <td><span class='badge warning'>⚠ Silent</span></td>\n"
        else:
            html_dashboard += f"                        <td><span class='badge success'>✓ Active</span></td>\n"
        
        html_dashboard += "                    </tr>\n"
    
    html_dashboard += """                </tbody>
            </table>
        </div>
        
        <!-- DECOMP TAB -->
        <div id="decomp" class="tab-content">
            <div class="stats">
                <div class="stat-card">
                    <div class="stat-label">Total Decomp Tasks</div>
                    <div class="stat-value">""" + str(len(decomp_taskers)) + """</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Active</div>
                    <div class="stat-value">""" + str(len(decomp_taskers) - len(decomp_silent)) + """</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Silent (48+hrs)</div>
                    <div class="stat-value" style="color: #dc2626;">""" + str(len(decomp_silent)) + """</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Total Decomp Items</div>
                    <div class="stat-value">""" + str(sum(sum(dates.values()) for dates in decomp_data.values())) + """</div>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Tasker Name</th>
"""
    
    for date_header in date_headers:
        html_dashboard += f"                        <th class='date'>{date_header}</th>\n"
    
    html_dashboard += """                        <th class='date'>Total</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    silent_decomp_names = {s["name"] for s in decomp_silent}
    
    for tasker in sorted(decomp_taskers):
        row_class = "silent" if tasker in silent_decomp_names else "active"
        html_dashboard += f"                    <tr class='{row_class}'>\n"
        html_dashboard += f"                        <td>{tasker}</td>\n"
        
        daily_counts = []
        for date in sorted_dates:
            count = decomp_data[tasker].get(date, 0)
            daily_counts.append(count)
            html_dashboard += f"                        <td class='date'>{count if count > 0 else '-'}</td>\n"
        
        total = sum(daily_counts)
        html_dashboard += f"                        <td class='date'>{total}</td>\n"
        
        if tasker in silent_decomp_names:
            html_dashboard += f"                        <td><span class='badge warning'>⚠ Silent</span></td>\n"
        else:
            html_dashboard += f"                        <td><span class='badge success'>✓ Active</span></td>\n"
        
        html_dashboard += "                    </tr>\n"
    
    html_dashboard += """                </tbody>
            </table>
        </div>
        
        <!-- SILENT TASKERS & NOTES TAB -->
        <div id="silent" class="tab-content">
            <h2 style="margin-bottom: 20px;">Silent Taskers & Outreach Notes</h2>
"""
    
    all_silent = sorted(claim_silent + decomp_silent, key=lambda x: x["hours_silent"], reverse=True)
    
    if all_silent:
        html_dashboard += """            <table>
                <thead>
                    <tr>
                        <th>Name</th>
                        <th>Last Activity</th>
                        <th>Silent For</th>
                        <th>Notes / Reason</th>
                        <th>Last Outreach</th>
                    </tr>
                </thead>
                <tbody>
"""
        for silent in all_silent:
            name = silent["name"]
            tasker_notes = notes.get(name, {})
            note_text = tasker_notes.get("reason", "No notes")
            last_outreach = tasker_notes.get("last_outreach", "Not reached out yet")
            
            html_dashboard += f"""                    <tr>
                        <td>{name}</td>
                        <td>{silent['last_task_date']}</td>
                        <td>{silent['days_silent']} days</td>
                        <td class="notes-column">{note_text}</td>
                        <td style="font-size: 12px; color: #666;">{last_outreach}</td>
                    </tr>
"""
        html_dashboard += """                </tbody>
            </table>
"""
    else:
        html_dashboard += """            <p style="color: #666; padding: 20px; text-align: center;">Everyone is active! No one is silent.</p>
"""
    
    html_dashboard += """        </div>
        
        <!-- HISTORIC DATA TAB -->
        <div id="historic" class="tab-content">
            <p style="color: #666; padding: 20px; text-align: center;">Historic data will appear here as you track more weeks.</p>
        </div>
    </div>
    
    <script>
        function openTab(evt, tabName) {
            var i, tabcontent, tabbuttons;
            tabcontent = document.getElementsByClassName("tab-content");
            for (i = 0; i < tabcontent.length; i++) {
                tabcontent[i].classList.remove("active");
            }
            tabbuttons = document.getElementsByClassName("tab-button");
            for (i = 0; i < tabbuttons.length; i++) {
                tabbuttons[i].classList.remove("active");
            }
            document.getElementById(tabName).classList.add("active");
            evt.currentTarget.classList.add("active");
        }
    </script>
</body>
</html>"""
    
    # Write dashboard
    dashboard_path = Path.cwd() / "dashboard.html"
    with open(dashboard_path, 'w') as f:
        f.write(html_dashboard)
    print(f"✓ Generated dashboard.html")
    
    # === APPROVAL UI with Slack API Call ===
    all_silent_merged = claim_silent + decomp_silent
    all_silent_merged.sort(key=lambda x: x["hours_silent"], reverse=True)
    
    approval_html = f"""<!DOCTYPE html>
<html>
<head>
    <title>Approval UI - Send Slack DMs</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 32px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        h1 {{ font-size: 28px; font-weight: 600; color: #1a1a1a; margin-bottom: 4px; }}
        .meta {{ color: #666; font-size: 14px; margin-bottom: 20px; }}
        .intro {{ background: #eff6ff; border-left: 4px solid #3b82f6; padding: 16px; border-radius: 4px; margin-bottom: 24px; color: #1e40af; font-size: 14px; line-height: 1.6; }}
        .person-card {{ border: 1px solid #e0e0e0; padding: 20px; margin-bottom: 16px; border-radius: 6px; background: #fafafa; }}
        .person-name {{ font-weight: 600; font-size: 16px; color: #1a1a1a; margin-bottom: 8px; }}
        .person-meta {{ color: #666; font-size: 13px; margin: 4px 0; }}
        .message-preview {{ background: #f0f0f0; padding: 12px; border-radius: 4px; margin: 12px 0; font-size: 13px; line-height: 1.5; color: #333; border-left: 3px solid #3b82f6; }}
        .notes-input {{ width: 100%; padding: 8px; border: 1px solid #ccc; border-radius: 4px; font-size: 13px; margin: 8px 0; }}
        .checkbox {{ display: flex; align-items: center; margin-top: 12px; }}
        .checkbox input {{ margin-right: 10px; width: 18px; height: 18px; cursor: pointer; }}
        .checkbox label {{ cursor: pointer; font-size: 14px; color: #1a1a1a; margin: 0; }}
        .controls {{ margin-top: 32px; padding-top: 24px; border-top: 1px solid #e0e0e0; display: flex; gap: 12px; }}
        button {{ padding: 10px 16px; font-size: 14px; border: 1px solid #ccc; border-radius: 4px; cursor: pointer; background: white; color: #1a1a1a; font-weight: 500; }}
        button:hover {{ background: #f5f5f5; }}
        .send-btn {{ background: #10b981; color: white; border-color: #10b981; }}
        .send-btn:hover {{ background: #059669; }}
        .no-people {{ color: #666; padding: 32px 20px; text-align: center; font-size: 15px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Slack Check-in & Notes</h1>
        <div class="meta">{datetime.now().strftime('%A, %B %d, %Y at %I:%M %p')} EST</div>
        <div class="intro">Review people who haven't worked in 48+ hours. Add notes on why they're silent, then send personalized Slack DMs.</div>
        
        <form id="approval-form">
"""
    
    if not all_silent_merged:
        approval_html += """            <div class="no-people">
                <div style="font-size: 32px; margin-bottom: 12px;">✅</div>
                <p><strong>Everyone is active!</strong></p>
                <p>No one needs a check-in message today.</p>
            </div>
"""
    else:
        for silent in all_silent_merged:
            name = silent["name"]
            first_name = silent["first_name"]
            tasker_notes = notes.get(name, {})
            existing_reason = tasker_notes.get("reason", "")
            
            approval_html += f"""            <div class="person-card">
                <p class="person-name">{name}</p>
                <p class="person-meta">Last activity: <strong>{silent['last_task_date']}</strong> ({silent['days_silent']} days ago)</p>
                
                <div class="message-preview">
                    <strong>Message preview:</strong><br>
                    Hey {first_name}, I noticed you haven't submitted any tasks in the past couple of days. If you're facing any blockers or have questions, let me know and I'm happy to help!
                </div>
                
                <label style="font-size: 13px; color: #666;">Why are they silent? (add reason/notes)</label>
                <textarea name="{name}_notes" class="notes-input" placeholder="e.g., Sick leave, Technical issues, Waiting for guidance..." rows="2">{existing_reason}</textarea>
                
                <div class="checkbox">
                    <input type="checkbox" id="{name}" name="selected_{name}" value="{name}">
                    <label for="{name}">Send Slack DM to {first_name}</label>
                </div>
            </div>
"""
    
    approval_html += """        </form>
        
        <div class="controls">
            <button onclick="selectAll()">Select All</button>
            <button onclick="clearAll()">Clear All</button>
            <button class="send-btn" onclick="sendMessages()">Send & Save Notes</button>
        </div>
    </div>
    
    <script>
        function selectAll() {
            document.querySelectorAll('#approval-form input[type="checkbox"]').forEach(cb => cb.checked = true);
        }
        
        function clearAll() {
            document.querySelectorAll('#approval-form input[type="checkbox"]').forEach(cb => cb.checked = false);
        }
        
        async function sendMessages() {
            const selected = [];
            
            document.querySelectorAll('#approval-form input[type="checkbox"]:checked').forEach(cb => {
                const taskName = cb.name.replace('selected_', '');
                const notesField = document.querySelector(`textarea[name="${taskName}_notes"]`);
                const reason = notesField ? notesField.value : "";
                
                selected.push({
                    name: taskName,
                    reason: reason
                });
            });
            
            if (selected.length === 0) {
                alert('Please select at least one person to message.');
                return;
            }
            
            let message = `Ready to send DMs to ${selected.length} person${selected.length !== 1 ? 's' : ''}?\\n\\n`;
            selected.forEach(s => {
                message += `• ${s.name}\\n`;
            });
            
            if (confirm(message)) {
                try {
                    const response = await fetch('/send_slack_messages', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({{ people: selected }})
                    });
                    
                    if (response.ok) {
                        alert('✓ Messages sent and notes saved!');
                    } else {
                        alert('✗ Error sending messages. Check console.');
                    }
                } catch (error) {
                    alert('✗ Error: ' + error.message);
                }
            }
        }
    </script>
</body>
</html>"""
    
    # Write approval UI
    approval_path = Path.cwd() / "approval_ui.html"
    with open(approval_path, 'w') as f:
        f.write(approval_html)
    print(f"✓ Generated approval_ui.html")

def main():
    print(f"\n=== Task Tracker Processor ===")
    print(f"Working directory: {Path.cwd()}")
    
    # Load Slack users
    slack_users = load_slack_users()
    if not slack_users:
        print("ERROR: Could not load Slack users")
        return
    print(f"✓ Loaded {len(slack_users)} Slack users")
    
    # Find Excel
    excel_path = find_latest_excel()
    if not excel_path:
        return
    print(f"✓ Found Excel: {excel_path.name}")
    
    # Read Claim Sheet
    claim_data, claim_taskers = read_claim_sheet(excel_path)
    if claim_data is None:
        return
    print(f"✓ Claim Sheet: {len(claim_taskers)} taskers")
    
    # Read Decomp Sheet
    decomp_data, decomp_taskers = read_decomp_sheet(excel_path)
    if decomp_data is None:
        decomp_data = {}
        decomp_taskers = []
    print(f"✓ Model Decomp: {len(decomp_taskers)} taskers")
    
    # Generate dashboards
    print(f"\n✓ Generating dashboards...")
    generate_dashboards(claim_data, claim_taskers, decomp_data, decomp_taskers, slack_users)
    
    print(f"\n✓ COMPLETE!")
    print(f"Dashboard: https://aashnagoel.github.io/task-tracker/dashboard.html")
    print(f"Approval UI: https://aashnagoel.github.io/task-tracker/approval_ui.html")

if __name__ == "__main__":
    main()
